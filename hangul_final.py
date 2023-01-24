import pprint
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
L_table = {
    0: 0.3,
    3: 0.6, 
    7: 0.7,
    15: 0.8,
    16: 0.9,
    20: 0.4,
    21: 0.8
}

def color(final):
    if final == '':
        return 'ffffff'
    jung = ord(final[1]) - 0x1161
    if jung >= 3:
        jung = jung - 1
    H = jung * 18
    L = 0.5
    if len(final) > 2:
        jong = ord(final[2]) - 0x11A8
        L = L_table[jong]
    k = divmod(H/30, 12)[1]
    f = (lambda n: L - min(L, 1-L) * max(-1, min(divmod(n + H/30, 12)[1]-3, 9-divmod(n + H/30, 12)[1], 1)))
    s = (lambda x: "%02x" % int(round(x * 255)))
    return s(f(0)) + s(f(8)) + s(f(4))

data = dict()
initals = '幫滂並明端透定泥知徹澄孃精清從心邪莊初崇生俟章昌常書船見溪羣疑影曉匣云以日來'

fA = '東冬鍾江支脂之微魚模虞泰廢夬佳皆祭齊咍灰'
fB = '眞臻諄痕魂欣文寒桓元刪山仙先豪肴宵蕭歌戈'
fC = '麻唐陽庚耕清青登蒸尤侯幽侵談嚴凡銜咸鹽添覃'
finals = fA + fB + fC

with open('result.csv') as rhymefile:
    rhymefile.readline()
    for line in rhymefile.readlines():
        row = str(line).rstrip().split(',')
        inital = row[1]
        final = row[2]
        if row[3] == '入':
            final = final + '入'
        sound = row[23]
        if final == '':
            continue
        if final not in data.keys():
            data[final] = []
            for i in range(len(initals)):
                data[final].append(dict())
        stat = data[final][initals.find(inital)]
        for val in sound.split(' '):
            if val != '':
                if val not in stat.keys():
                    stat[val] = 0
                stat[val] += 1

pprint.pprint(data)

for final in data.keys():
    for initalidx in range(len(initals)):
        stat = data[final][initalidx]
        maxcount = 0
        maxchar = ''
        for char in stat.keys():
            if stat[char] > maxcount:
                maxcount = stat[char]
                maxchar = char
        data[final][initalidx] = maxchar
        
def setborder(ws, cols, i):
    b = Border(right=Side(border_style='thin', color='000000'))
    rnum = str(i + 2)
    for col in cols:
        ws[col + rnum].border = b

wb = Workbook()
sheet = wb.active

sheet.append(['韻'] + [c for c in initals])
for i, final in enumerate(sorted(data.keys(), key=lambda x: finals.find(x[0]))):
    sheet.append([final] + data[final])
    for initalidx in range(len(initals)):
        c = color(data[final][initalidx])
        pfill = PatternFill(start_color=c, end_color=c, fill_type='solid')
        sheet.cell(row=i+2, column=initalidx+2).fill = pfill
        setborder(sheet, ['E', 'I', 'M', 'R', 'W', 'AB', 'AF', 'AK'], i)

for i in range(len(initals)):
    sheet.column_dimensions[get_column_letter(i + 2)].width = 2

sheet.freeze_panes = sheet['B2']
wb.save('hangul_final.xlsx')