
import csv
import unicodedata
import hashlib
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def sep_mandarin(mandarin):
    if mandarin == '':
        return ('', '')
    umap = {
        'u' : 'ü',
        'ū' : 'ǖ',
        'ú' : 'ǘ',
        'ǔ' : 'ǚ',
        'ù' : 'ǜ'
    }
    if mandarin[0] == 'w':
        if mandarin[1] in 'uūúǔù':
            return ('', mandarin[1:])
        else:
            return ('', 'u' + mandarin[1:])
    if mandarin[0] == 'y':
        if mandarin[1] in 'iīíǐì':
            return ('', mandarin[1:])
        elif mandarin[1] in 'uūúǔù':
            return ('', umap[mandarin[1]] + mandarin[2:])
        else:
            return ('', 'i' + mandarin[1:])
    if mandarin[0] in 'jqx' and mandarin[1] in 'uūúǔù':
        return (mandarin[0], umap[mandarin[1]] + mandarin[2:])
    index = 0
    for ch in mandarin:
        if ch in 'aāáǎàeēéěèiīíǐìoōóǒòuūúǔùüǖǘǚǜ':
            break
        index = index + 1
    
    return (mandarin[:index], mandarin[index:])

def sep_cantonese(cantonese):
    if cantonese == '':
        return ('', '')
    if cantonese == 'm' or cantonese == 'ng':
        return ('', cantonese)
    if cantonese[0] == 'j' and (cantonese[1] == 'i' or cantonese[1] == 'y'):
        return ('', cantonese[1:])
    if cantonese[0] == 'w' and cantonese[1] == 'u':
        return ('', cantonese[1:])
    index = 0
    for ch in cantonese:
        if ch in 'aeiouy':
            break
        index = index + 1
    
    return (cantonese[:index], cantonese[index:])

def sep_japanese(japanese):
    if japanese == '':
        return ('', '')
    if len(japanese) >= 2:
        if japanese[1] == 'h':
            if japanese[0] == 's':
                if japanese[2] == 'i':
                    return ('s', japanese[2:])
                else:
                    return ('s', 'y' + japanese[2:])
            elif japanese[0] == 'c':
                if japanese[2] == 'i':
                    return ('t', japanese[2:])
                else:
                    return ('t', 'y' + japanese[2:])
        elif japanese[0] == 't' and japanese[1] == 's':
            return ('t', japanese[2:])
    if japanese[0] == 'j':
        if japanese[1] == 'i':
            return ('z', japanese[1:])
        else:
            return ('z', 'y' + japanese[1:])
    elif japanese[0] == 'f':
        return ('h', japanese[1:])
    elif japanese[0] in 'kstnhmrgzdbp':
        return (japanese[:1], japanese[1:])
    return ('', japanese)

def sep_vietnamese(vietnamese):
    if vietnamese == '':
        return ('', '')
    if len(vietnamese) >= 2:
        if len(vietnamese) >= 3 and vietnamese[2] == 'h':
            return ('ng', vietnamese[3:])
        if len(vietnamese) >= 2 and vietnamese[0] == 'g':
            if vietnamese[1] == 'h':
                return ('g', vietnamese[2:])
            elif vietnamese[1] == 'i':
                return (vietnamese[:2], vietnamese[2:])
            elif vietnamese[1] == 'ị':
                return ('gi', vietnamese[1:])
    if vietnamese[0] == 'q' or vietnamese[0] == 'k':
        return ('c', vietnamese[1:])
    elif vietnamese[0] in 'bcdđghlmnprstvx':
        return (vietnamese[:1], vietnamese[1:])
    return ('', vietnamese)

def sep_korean(korean):
    if korean == '':
        return ('', '')
    nfd = unicodedata.normalize('NFD', korean)
    return (nfd[0], unicodedata.normalize('NFC', chr(0x110B) + nfd[1:]))
    
unidata = dict()

def myzip(dual):
    first = ''
    second = ''
    for (inital, final) in dual:
        if inital == '' and final != '':
            first = first + ' ' + '/'
        else:
            first = first + ' ' + inital
        second = second + ' ' + final
    return (first[1:], second[1:])

def additemcount(d, s):
    for item in s.split(' '):
        if item == '':
            continue
        if item not in d.keys():
            d[item] = 1
        else:
            d[item] = d[item] + 1

def initcount():
    return ((dict(), dict()), (dict(), dict()), (dict(), dict()), (dict(), dict()), (dict(), dict()), (dict(), dict()))

with open('Unihan_Readings.txt') as unireadings:
    for line in unireadings.readlines():
        row = line.rstrip().split('\t')
        dataindex = chr(int(row[0][2:], 16))
        if dataindex not in unidata.keys():
            unidata[dataindex] = dict()
        unidata[dataindex][row[1]] = row[2]

with open('Unihan_DictionaryLikeData.txt') as unidic:
    for line in unidic.readlines():
        row = line.rstrip().split('\t')
        dataindex = chr(int(row[0][2:], 16))
        if dataindex not in unidata.keys():
            unidata[dataindex] = dict()
        unidata[dataindex][row[1]] = row[2]

with open('kw.csv', newline='') as csvfile:
    kwreader = csv.reader(csvfile, delimiter=' ')
    for row in kwreader:
        row = str(row).split(',')
        kanji = row[6]
        if kanji not in unidata.keys():
            unidata[kanji] = dict()
        final = row[3][1:-1][::-1]
        try:
            unidata[kanji]['initial'] = row[3][0]
            unidata[kanji]['final'] = final
            unidata[kanji]['tone'] = row[3][-1]
            if row[5] != '':
                unidata[kanji]['upper'] = row[5][0]
                unidata[kanji]['lower'] = row[5][1]
            elif row[4] != '':
                unidata[kanji]['upper'] = row[4][0]
                unidata[kanji]['lower'] = row[4][1]
            else:
                unidata[kanji]['upper'] = ''
                unidata[kanji]['lower'] = ''
        except:
            print(row)

rhymes = dict()
with open('rhyme.csv') as rhymefile:
    for line in rhymefile.readlines():
        row = str(line).rstrip().split(',')
        rhymes[row[4]] = dict()
        rhymes[row[4]]['she'] = row[0]
        rhymes[row[4]]['yun'] = row[1]
        rhymes[row[4]]['deng'] = row[2]
        rhymes[row[4]]['hu'] = row[3]

mid = []

for k, v in unidata.items():
    man = v.get('kHanyuPinyin', '')
    if man != '':
        man = man.split(':')[1].replace(',', ' ')
        
    output = [k, v.get('initial', ''), v.get('final', ''), v.get('tone', ''),
    v.get('upper', ''), v.get('lower', ''), man,
    v.get('kCantonese', ''), v.get('kJapaneseOn', '').lower(), v.get('kVietnamese', ''),
    v.get('kHangul', '').translate(str.maketrans({'0':'', '1':'', '2':'', '3':'', 'E':'', 'N':'', 'X':'', ':':''})),
    v.get('kPhonetic', '').replace('*', ''), v.get('kFenn', '').split(' ')[0]]
    mid.append(output)

wb = Workbook()

# data table

datatable = wb.active
datatable.title = 'datatable'

soothilltable = wb.create_sheet('soothill')
pictotable = wb.create_sheet('pictolist')

sd = dict()
pd = dict()
initiald = dict()
finald = dict()

datatable.append(['unicode', 'kanji', 'initial', 'final', 'tone',
        'upper', 'lower', 'she', 'yun', 'division', 'rounding',
        'Minitial', 'Mfinal', 'Mandarin', 'Cinitial', 'Cfinal', 'Cantonese',
        'Jinitial', 'Jfinal', 'Japanese', 'Vinitial', 'Vfinal', 'Vietnamese',
        'Kinitial', 'Kfinal', 'Korean', 'Pictophonetic', 'Soothill', 'Frequency'])

soothilltable.append(['number', 'kanjis',
                      'initial', 'final',
                      'Mini', 'Mfin',
                      'Cini', 'Cfin',
                      'Jini', 'Jfin',
                      'Vini', 'Vfin',
                      'Kini', 'Kfin'])

pictotable.append(['number', 'kanjis',
                      'initial', 'final',
                      'Mini', 'Mfin',
                      'Cini', 'Cfin',
                      'Jini', 'Jfin',
                      'Vini', 'Vfin',
                      'Kini', 'Kfin'])

for data in sorted(mid, key=lambda x: x[0]):
    man = myzip([sep_mandarin(x) for x in data[6].split(' ')])
    can = myzip([sep_mandarin(x) for x in data[7].split(' ')])
    jap = myzip([sep_japanese(x) for x in data[8].split(' ')])
    vet = myzip([sep_vietnamese(x) for x in data[9].split(' ')])
    kor = myzip([sep_korean(x) for x in data[10].split(' ')])

    def inccount(count):
        additemcount(count[0][0], data[1])
        additemcount(count[0][1], data[2])
        additemcount(count[1][0], man[0])
        additemcount(count[1][1], man[1])
        additemcount(count[2][0], can[0])
        additemcount(count[2][1], can[1])
        additemcount(count[3][0], jap[0])
        additemcount(count[3][1], jap[1])
        additemcount(count[4][0], vet[0])
        additemcount(count[4][1], vet[1])
        additemcount(count[5][0], kor[0])
        additemcount(count[5][1], kor[1])
    
    if data[2] != '':
        ylist = [
            rhymes[data[2]]['she'],
            rhymes[data[2]]['yun'],
            rhymes[data[2]]['deng'],
            rhymes[data[2]]['hu'],
        ]
    else:
        ylist = ['', '', '', '']
    dlist = [
        man[0], man[1], data[6],
        can[0], can[1], data[7],
        jap[0], jap[1], data[8],
        vet[0], vet[1], data[9],
        kor[0], kor[1], data[10]
    ]
    if data[12] != '':
        flist = [data[12][:-1], data[12][-1]]
        if flist[0][-1] == 'a':
            flist[0] = flist[0][:-1]
        sn = int(flist[0])
        if sn not in sd.keys():
            sd[sn] = ['', initcount()]
        sd[sn][0] += data[0]
        inccount(sd[sn][1])
    else:
        flist = ['', '']

    datatable.append(['u+'+hex(ord(data[0]))[2:]] + data[0:6] + ylist + dlist + [data[11]] + flist)

    for pnraw in data[11].split():
        pn = int(pnraw[:-1] if pnraw[-1].isalpha() else pnraw)
        if pn not in pd.keys():
            pd[pn] = ['', initcount()]
        pd[pn][0] += data[0]
        inccount(pd[pn][1])

    if kor[0] != '':
        data[1] = kor[0][0]
        data[2] = kor[1].split()[0]
        data[3] = ''
    else:
        data[2] = ''
    
    if data[2] != '':
        finalwithtone = data[2] + data[3]
        if finalwithtone not in finald.keys():
            finald[finalwithtone] = dict()
        if data[1] not in finald[finalwithtone].keys():
            finald[finalwithtone][data[1]] = initcount()
        inccount(finald[finalwithtone][data[1]])

        if data[2] not in initiald.keys():
            initiald[data[2]] = dict() 
        if data[1] not in initiald[data[2]].keys():
            initiald[data[2]][data[1]] = initcount()
        inccount(initiald[data[2]][data[1]])
    
    


def strfy(d):
    s = ''
    for k, v in dict(sorted(d.items(), key=lambda item: item[1], reverse=True)).items():
        s += k + ':' + str(v) + ' '
    return s

def countl(count):
    return [strfy(count[0][0]), strfy(count[0][1]),
            strfy(count[1][0]), strfy(count[1][1]),
            strfy(count[2][0]), strfy(count[2][1]),
            strfy(count[3][0]), strfy(count[3][1]),
            strfy(count[4][0]), strfy(count[4][1]),
            strfy(count[5][0]), strfy(count[5][1])
    ]

def setborder(ws, cols, i):
    b = Border(right=Side(border_style='thin', color='000000'))
    rnum = str(i + 2)
    for col in cols:
        ws[col + rnum].border = b

for k, v in dict(sorted(sd.items())).items():
    soothilltable.append([k, v[0]] + countl(v[1]))

for k, v in dict(sorted(pd.items())).items():
    pictotable.append([k, v[0]] + countl(v[1]))

inittables = [
    wb.create_sheet('Minitlist'),
    wb.create_sheet('Cinitlist'),
    wb.create_sheet('Jinitlist'),
    wb.create_sheet('Vinitlist'),
    wb.create_sheet('Kinitlist')
]

finaltables = [
    wb.create_sheet('Mfinallist'),
    wb.create_sheet('Cfinallist'),
    wb.create_sheet('Jfinallist'),
    wb.create_sheet('Vfinallist'),
    wb.create_sheet('Kfinallist')
]

#initials = '幫滂並明端透定泥知徹澄孃精清從心邪莊初崇生俟章昌常書船見溪羣疑影匣云曉來日以'
#finals = '東冬鍾江支脂之微魚模虞泰廢夬佳皆祭齊咍灰眞臻諄痕魂欣文元寒桓刪山仙先豪肴宵蕭歌戈麻唐陽庚耕清青登蒸尤侯幽侵談嚴凡銜咸鹽添覃'
initials = 'ᄋᄀᄏᄒᄂᄃᄐᄅᄆᄇᄑᄉᄌᄎ'
finals = '아악안알암압앙애액앵야약양어억언얼엄업엉에엔여역연열염엽영예오옥온올옹옺와왁완왈왕왜외왹욍요욕용우욱운울움웅원월웨위유육윤율융윽은을음읍응의이익인일임입잉'

for inittable in inittables:
    #inittable.append(['韻', '攝', '等'] + [c for c in initials])
    inittable.append(['韻'] + [c for c in initials])
    for i in range(len(initials)+2):
        inittable.column_dimensions[get_column_letter(i + 2)].width = 4

for finaltable in finaltables:
    #finaltable.append(['韻', '攝', '等'] + [c for c in initials])
    finaltable.append(['韻'] + [c for c in initials])
    for i in range(len(initials)+2):
        finaltable.column_dimensions[get_column_letter(i + 2)].width = 4

remove_tone = {
    'ā': 'a', 'á': 'a', 'ǎ': 'a', 'à': 'a',
    'ē': 'e', 'é': 'e', 'ě': 'e', 'è': 'e',
    'ī': 'i', 'í': 'i', 'ǐ': 'i', 'ì': 'i',
    'ō': 'o', 'ó': 'o', 'ǒ': 'o', 'ò': 'o',
    'ū': 'u', 'ú': 'u', 'ǔ': 'u', 'ù': 'u',
    'ǖ': 'ü', 'ǘ': 'ü', 'ǚ': 'ü', 'ǜ': 'ü'
}
def getmaxfill(x):
    if len(x) == 0:
        c = 'ffffff'
    else:
        maxkey = max(x,key=x.get)
        if maxkey[0] in remove_tone:
            maxkey = remove_tone[maxkey[0]] + maxkey[1:]
        if len(maxkey) >= 2 and maxkey[1] in remove_tone:
            maxkey = maxkey[0] + remove_tone[maxkey[1]] + maxkey[2:]
        if maxkey[-1] in '123456':
            maxkey = maxkey[:-1]
        m = hashlib.sha1()
        m.update(maxkey.encode('utf-8'))
        c = m.hexdigest()[0:6]
    pfill = PatternFill(start_color=c, end_color=c, fill_type='solid')
    return pfill

L_table = {
    0: 0.3,
    3: 0.6, 
    7: 0.7,
    15: 0.8,
    16: 0.9,
    20: 0.4,
    21: 0.8
}

def hangul_maxfill(x):

    if len(x) == 0:
        c = 'ffffff'
    else:
        final = max(x,key=x.get)
        final = unicodedata.normalize('NFD', final)
        if final == '':
            return 'ffffff'
        jung = ord(final[1]) - 0x1161
        if jung >= 3:
            jung = jung - 1
        H = jung * 18
        L = 0.5
        if len(final) > 2:
            jong = ord(final[2]) - 0x11A8
            L = L_table[jong]
        k = divmod(H/30, 12)[1]
        f = (lambda n: L - min(L, 1-L) * max(-1, min(divmod(n + H/30, 12)[1]-3, 9-divmod(n + H/30, 12)[1], 1)))
        s = (lambda x: "%02x" % int(round(x * 255)))
        c = s(f(0)) + s(f(8)) + s(f(4))
    pfill = PatternFill(start_color=c, end_color=c, fill_type='solid')
    return pfill

def evalchar(x):
    if x == '入':
        return 0
    elif x == '平':
        return 1
    elif x == '去':
        return 2
    elif x == '上':
        return 3
    else:
        return ord(x)

def finalsortkey(x):
    r = finals.find(x[0]) * (2 ** 54)
    if len(x) >= 2:
        r += evalchar(x[1]) * (2 ** 36)
        if len(x) >= 3:
            r += evalchar(x[2]) * (2 ** 18)
            if len(x) >= 4:
                r += evalchar(x[3])
    return r

def remove_extra(x):
    return x.rstrip('平上去入')

for langno in range(5):
    for idx, final in enumerate(sorted(finald.keys(), key=finalsortkey)):
        finaldict = finald[final]
        fincols = []
        for jdx, initial in enumerate(initials):
            if initial not in finaldict.keys():
                fincols.append('')
            else:
                langdicts = finaldict[initial][langno+1]
                fincols.append(strfy(langdicts[1]))
        #finaltables[langno].append([final, rhymes[remove_extra(final)]['she'], rhymes[remove_extra(final)]['deng']] + fincols)
        finaltables[langno].append([final] + fincols)

        for jdx, initial in enumerate(initials):
            if initial in finaldict.keys():
                langdicts = finaldict[initial][langno+1]
                finaltables[langno].cell(row=idx+2, column=jdx+2).fill = \
                    getmaxfill(langdicts[1]) if langno != 4 else hangul_maxfill(langdicts[1])


    for idx, final in enumerate(sorted(initiald.keys(), key=finalsortkey)):
        finaldict = initiald[final]
        inicols = []
        for jdx, initial in enumerate(initials):
            if initial not in finaldict.keys():
                inicols.append('')
            else:
                langdicts = finaldict[initial][langno+1]
                inicols.append(strfy(langdicts[0]))
        #inittables[langno].append([final, rhymes[final]['she'], rhymes[final]['deng']] + inicols)
        inittables[langno].append([final] + inicols)


        for jdx, initial in enumerate(initials):
            if initial in finaldict.keys():
                langdicts = finaldict[initial][langno+1]
                inittables[langno].cell(row=idx+2, column=jdx+2).fill = getmaxfill(langdicts[0])

wb.save('result-kor.xlsx')
